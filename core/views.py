import calendar
import datetime
import io
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from .forms import ExpenseCategoryForm, OwnerDeductionForm, UnitForm
from .models import (
    Distribution, ExpenseCategory, MonthlyExpense, Owner,
    OwnerDeduction, RentalIncome, Unit,
)

# ─── helpers ────────────────────────────────────────────────────────────────

def _month_from_request(request) -> datetime.date:
    """Return the first day of the month from GET params; default = this month."""
    today = datetime.date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12 and 2000 <= year <= 2100):
            raise ValueError
        return datetime.date(year, month, 1)
    except (ValueError, TypeError):
        return datetime.date(today.year, today.month, 1)


def _adjacent_month(d: datetime.date, delta: int) -> datetime.date:
    """Shift d by delta months (±)."""
    month = d.month + delta
    year = d.year + (month - 1) // 12
    month = (month - 1) % 12 + 1
    return datetime.date(year, month, 1)


def _month_summary(month_date: datetime.date) -> dict:
    total_income = (
        RentalIncome.objects.filter(month=month_date)
        .aggregate(v=Sum('amount'))['v'] or Decimal('0')
    )
    total_expense = (
        MonthlyExpense.objects.filter(month=month_date)
        .aggregate(v=Sum('amount'))['v'] or Decimal('0')
    )
    net = total_income - total_expense
    num_owners = Owner.objects.count() or 4
    per_person = (net / num_owners).quantize(Decimal('0.01'))
    return {
        'total_income': total_income,
        'total_expense': total_expense,
        'net': net,
        'per_person': per_person,
        'num_owners': num_owners,
    }


def _nav_context(month_date: datetime.date) -> dict:
    return {
        'month_date': month_date,
        'prev_month': _adjacent_month(month_date, -1),
        'next_month': _adjacent_month(month_date, 1),
    }


# ─── dashboard ──────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    month_date = _month_from_request(request)
    summary = _month_summary(month_date)

    # Last 6 months for quick navigation
    recent_months = [_adjacent_month(month_date, -i) for i in range(5, -1, -1)]

    # Rental income snapshot
    income_qs = RentalIncome.objects.filter(month=month_date).select_related('unit')

    # Expense snapshot
    expense_qs = MonthlyExpense.objects.filter(month=month_date).select_related('category')

    # Distribution snapshot
    dist_qs = Distribution.objects.filter(month=month_date).select_related('owner')

    ctx = {
        **_nav_context(month_date),
        **summary,
        'recent_months': recent_months,
        'incomes': income_qs,
        'expenses': expense_qs,
        'distributions': dist_qs,
    }
    return render(request, 'core/dashboard.html', ctx)


# ─── rental income ───────────────────────────────────────────────────────────

@login_required
def rental_income(request):
    month_date = _month_from_request(request)
    units = Unit.objects.filter(is_active=True)

    income_map = {
        ri.unit_id: ri
        for ri in RentalIncome.objects.filter(month=month_date)
    }

    rows = [{'unit': u, 'income': income_map.get(u.id)} for u in units]
    total = sum(ri.amount for ri in income_map.values())

    ctx = {
        **_nav_context(month_date),
        'rows': rows,
        'total': total,
    }
    return render(request, 'core/rental_income.html', ctx)


@login_required
def rental_income_save(request):
    if request.method != 'POST':
        return redirect('rental_income')

    try:
        month_date = datetime.date.fromisoformat(request.POST.get('month', ''))
    except ValueError:
        messages.error(request, 'Invalid month value.')
        return redirect('rental_income')

    for unit in Unit.objects.filter(is_active=True):
        raw = request.POST.get(f'amount_{unit.id}', '').strip()
        if raw == '':
            continue
        try:
            amount = Decimal(raw)
        except InvalidOperation:
            messages.warning(request, f'Invalid amount for {unit.name} — skipped.')
            continue
        notes = request.POST.get(f'notes_{unit.id}', '').strip()
        is_paid = request.POST.get(f'paid_{unit.id}') == 'on'
        RentalIncome.objects.update_or_create(
            unit=unit, month=month_date,
            defaults={'amount': amount, 'notes': notes, 'is_paid': is_paid},
        )

    messages.success(request, f'Rental income for {month_date.strftime("%B %Y")} saved.')
    return redirect(f'/rental-income/?year={month_date.year}&month={month_date.month}')


# ─── expenses ────────────────────────────────────────────────────────────────

@login_required
def expenses(request):
    month_date = _month_from_request(request)
    categories = ExpenseCategory.objects.all()

    expense_map = {
        e.category_id: e
        for e in MonthlyExpense.objects.filter(month=month_date).select_related('category')
    }

    rows = [{'category': c, 'expense': expense_map.get(c.id)} for c in categories]
    total = sum(e.amount for e in expense_map.values())

    ctx = {
        **_nav_context(month_date),
        'rows': rows,
        'total': total,
    }
    return render(request, 'core/expenses.html', ctx)


@login_required
def expenses_save(request):
    if request.method != 'POST':
        return redirect('expenses')

    try:
        month_date = datetime.date.fromisoformat(request.POST.get('month', ''))
    except ValueError:
        messages.error(request, 'Invalid month value.')
        return redirect('expenses')

    for cat in ExpenseCategory.objects.all():
        raw = request.POST.get(f'amount_{cat.id}', '').strip()
        notes = request.POST.get(f'notes_{cat.id}', '').strip()
        if raw == '':
            # Remove existing record if user cleared it
            MonthlyExpense.objects.filter(category=cat, month=month_date).delete()
            continue
        try:
            amount = Decimal(raw)
        except InvalidOperation:
            messages.warning(request, f'Invalid amount for {cat.name} — skipped.')
            continue
        MonthlyExpense.objects.update_or_create(
            category=cat, month=month_date,
            defaults={'amount': amount, 'notes': notes},
        )

    messages.success(request, f'Expenses for {month_date.strftime("%B %Y")} saved.')
    return redirect(f'/expenses/?year={month_date.year}&month={month_date.month}')


# ─── distribution ────────────────────────────────────────────────────────────

@login_required
def distribution(request):
    month_date = _month_from_request(request)
    summary = _month_summary(month_date)

    owners = Owner.objects.all()

    # Deductions per owner
    deduction_qs = OwnerDeduction.objects.filter(month=month_date).select_related('owner')
    deduction_map: dict[int, list] = {}
    for d in deduction_qs:
        deduction_map.setdefault(d.owner_id, []).append(d)

    # Saved distribution records
    dist_map = {
        d.owner_id: d
        for d in Distribution.objects.filter(month=month_date)
    }

    rows = []
    for owner in owners:
        owner_deductions = deduction_map.get(owner.id, [])
        total_ded = sum(d.amount for d in owner_deductions)
        net = summary['per_person'] - total_ded
        rows.append({
            'owner': owner,
            'gross': summary['per_person'],
            'deduction_list': owner_deductions,
            'total_deductions': total_ded,
            'net': net,
            'saved': dist_map.get(owner.id),
        })

    ctx = {
        **_nav_context(month_date),
        **summary,
        'rows': rows,
    }
    return render(request, 'core/distribution.html', ctx)


@login_required
def distribution_save(request):
    if request.method != 'POST':
        return redirect('distribution')

    try:
        month_date = datetime.date.fromisoformat(request.POST.get('month', ''))
    except ValueError:
        messages.error(request, 'Invalid month value.')
        return redirect('distribution')

    summary = _month_summary(month_date)

    for owner in Owner.objects.all():
        deductions_total = (
            OwnerDeduction.objects.filter(owner=owner, month=month_date)
            .aggregate(v=Sum('amount'))['v'] or Decimal('0')
        )
        net = summary['per_person'] - deductions_total
        notes = request.POST.get(f'notes_{owner.id}', '').strip()
        Distribution.objects.update_or_create(
            owner=owner, month=month_date,
            defaults={
                'gross_amount': summary['per_person'],
                'total_deductions': deductions_total,
                'net_amount': net,
                'notes': notes,
            },
        )

    messages.success(request, f'Distribution for {month_date.strftime("%B %Y")} saved.')
    return redirect(f'/distribution/?year={month_date.year}&month={month_date.month}')


# ─── deductions ──────────────────────────────────────────────────────────────

@login_required
def deduction_add(request):
    month_date = _month_from_request(request)

    if request.method == 'POST':
        form = OwnerDeductionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Deduction added.')
            d = form.instance
            return redirect(
                f'/distribution/?year={d.month.year}&month={d.month.month}'
            )
    else:
        form = OwnerDeductionForm(initial={'month': month_date})

    return render(request, 'core/deduction_form.html', {
        **_nav_context(month_date),
        'form': form,
        'title': 'Add Deduction',
    })


@login_required
def deduction_delete(request, pk):
    ded = get_object_or_404(OwnerDeduction, pk=pk)
    month = ded.month
    if request.method == 'POST':
        ded.delete()
        messages.success(request, 'Deduction removed.')
    return redirect(f'/distribution/?year={month.year}&month={month.month}')


# ─── expense categories ───────────────────────────────────────────────────────

@login_required
def expense_categories(request):
    cats = ExpenseCategory.objects.all()
    return render(request, 'core/expense_categories.html', {'categories': cats})


@login_required
def expense_category_create(request):
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category added.')
            return redirect('expense_categories')
    else:
        form = ExpenseCategoryForm()
    return render(request, 'core/expense_category_form.html', {'form': form, 'title': 'Add Expense Category'})


@login_required
def expense_category_edit(request, pk):
    cat = get_object_or_404(ExpenseCategory, pk=pk)
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST, instance=cat)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated.')
            return redirect('expense_categories')
    else:
        form = ExpenseCategoryForm(instance=cat)
    return render(request, 'core/expense_category_form.html', {
        'form': form, 'title': 'Edit Category', 'category': cat,
    })


@login_required
def expense_category_delete(request, pk):
    cat = get_object_or_404(ExpenseCategory, pk=pk)
    if request.method == 'POST':
        cat.delete()
        messages.success(request, 'Category deleted.')
        return redirect('expense_categories')
    return render(request, 'core/expense_category_confirm_delete.html', {'category': cat})


# ─── units ───────────────────────────────────────────────────────────────────

@login_required
def units(request):
    units_qs = Unit.objects.all()
    return render(request, 'core/units.html', {'units': units_qs})


@login_required
def unit_create(request):
    if request.method == 'POST':
        form = UnitForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unit created.')
            return redirect('units')
    else:
        form = UnitForm()
    return render(request, 'core/unit_form.html', {'form': form, 'title': 'Add Unit'})


@login_required
def unit_edit(request, pk):
    unit = get_object_or_404(Unit, pk=pk)
    if request.method == 'POST':
        form = UnitForm(request.POST, instance=unit)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unit updated.')
            return redirect('units')
    else:
        form = UnitForm(instance=unit)
    return render(request, 'core/unit_form.html', {
        'form': form, 'title': 'Edit Unit', 'unit': unit,
    })


# ─── report helpers ──────────────────────────────────────────────────────────

FLOOR_ORDER = ['7th_roof', '6th', '5th', '4th', '3rd', '2nd', '1st', 'garage']
SINGLE_FLOORS = {'7th_roof', 'garage'}

# Positions treated as "left" (first column) vs "right" (second column)
LEFT_POSITIONS = {'front', 'side1', 'roof', 'full', 'garage'}


def _build_income_rows(month_date):
    """
    Returns a list of dicts representing the paired-unit layout used in the Excel.
    Each row has: left_unit, left_amount, right_unit, right_amount, subtotal
    """
    income_map = {
        ri.unit_id: ri.amount
        for ri in RentalIncome.objects.filter(month=month_date).select_related('unit')
    }

    # Group active units by floor
    floor_groups: dict[str, list] = {}
    for unit in Unit.objects.filter(is_active=True).order_by('floor', 'position', 'name'):
        floor_groups.setdefault(unit.floor, []).append(unit)

    rows = []
    for floor in FLOOR_ORDER:
        group = floor_groups.get(floor, [])
        if not group:
            continue

        if floor in SINGLE_FLOORS:
            # Single-column row
            u = group[0]
            amt = income_map.get(u.id, Decimal('0'))
            rows.append({
                'left_name': u.name, 'left_amount': amt,
                'right_name': '', 'right_amount': None,
                'subtotal': amt,
                'is_single': True,
            })
            for u in group[1:]:
                amt = income_map.get(u.id, Decimal('0'))
                rows.append({
                    'left_name': u.name, 'left_amount': amt,
                    'right_name': '', 'right_amount': None,
                    'subtotal': amt,
                    'is_single': True,
                })
        else:
            # Pair units: try front/left vs back/right
            left = [u for u in group if u.position in LEFT_POSITIONS]
            right = [u for u in group if u.position not in LEFT_POSITIONS]

            # If both sides missing, treat all as left singles
            if not right:
                for u in left:
                    amt = income_map.get(u.id, Decimal('0'))
                    rows.append({
                        'left_name': u.name, 'left_amount': amt,
                        'right_name': '', 'right_amount': None,
                        'subtotal': amt,
                        'is_single': True,
                    })
            else:
                # Zip pairs; if uneven, extras appear as singles
                for l_unit, r_unit in zip(left, right):
                    la = income_map.get(l_unit.id, Decimal('0'))
                    ra = income_map.get(r_unit.id, Decimal('0'))
                    rows.append({
                        'left_name': l_unit.name, 'left_amount': la,
                        'right_name': r_unit.name, 'right_amount': ra,
                        'subtotal': la + ra,
                        'is_single': False,
                    })
                for u in left[len(right):] + right[len(left):]:
                    amt = income_map.get(u.id, Decimal('0'))
                    rows.append({
                        'left_name': u.name, 'left_amount': amt,
                        'right_name': '', 'right_amount': None,
                        'subtotal': amt,
                        'is_single': True,
                    })
    return rows


def _build_report_context(month_date):
    summary = _month_summary(month_date)

    income_rows = _build_income_rows(month_date)
    total_income = sum(r['subtotal'] for r in income_rows)

    expense_qs = MonthlyExpense.objects.filter(month=month_date).select_related('category')
    total_expense = sum(e.amount for e in expense_qs)

    owners = Owner.objects.all()
    deduction_qs = OwnerDeduction.objects.filter(month=month_date).select_related('owner')
    deduction_map: dict[int, list] = {}
    for d in deduction_qs:
        deduction_map.setdefault(d.owner_id, []).append(d)

    per_person = ((total_income - total_expense) / (owners.count() or 4)).quantize(Decimal('0.01'))

    dist_rows = []
    for owner in owners:
        owner_deds = deduction_map.get(owner.id, [])
        total_ded = sum(d.amount for d in owner_deds)
        net = per_person - total_ded
        dist_rows.append({
            'owner': owner,
            'gross': per_person,
            'deductions': owner_deds,
            'total_deductions': total_ded,
            'net': net,
        })

    return {
        **_nav_context(month_date),
        'income_rows': income_rows,
        'total_income': total_income,
        'expenses': expense_qs,
        'total_expense': total_expense,
        'net': total_income - total_expense,
        'dist_rows': dist_rows,
        'per_person': per_person,
        'num_owners': owners.count(),
    }


# ─── report views ─────────────────────────────────────────────────────────────

@login_required
def report(request):
    month_date = _month_from_request(request)
    ctx = _build_report_context(month_date)
    return render(request, 'core/report.html', ctx)


@login_required
def report_xlsx(request):  # noqa: PLR0912,PLR0915
    """Generate Excel report matching the original April_Property_House style."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    month_date = _month_from_request(request)
    ctx = _build_report_context(month_date)
    month_label = month_date.strftime('%B, %Y')

    wb = openpyxl.Workbook()

    # colours from original file
    STEEL = PatternFill('solid', fgColor='76A5AF')
    LGREEN = PatternFill('solid', fgColor='B6D7A8')
    LGREY = PatternFill('solid', fgColor='EFEFEF')
    DIST_HD = PatternFill('solid', fgColor='A2C4C9')
    GREEN_ROW = PatternFill('solid', fgColor='6AA84F')
    RED_ROW = PatternFill('solid', fgColor='E06666')

    BOLD = Font(bold=True)
    NORM = Font()
    TITLE = Font(bold=True, size=14)
    HDR_FONT = Font(bold=True, size=12)

    L = Alignment(horizontal='left', vertical='center', wrap_text=True)
    R = Alignment(horizontal='right', vertical='center')
    C = Alignment(horizontal='center', vertical='center')

    def _bdr(left='thin'):
        lside = Side(style='medium') if left == 'medium' else Side(style='thin')
        t = Side(style='thin')
        return Border(left=lside, right=t, top=t, bottom=t)

    BM = _bdr('medium')
    BT = _bdr('thin')

    def sc(ws, coord, value, font=None, fill=None, align=L, border=BT, fmt=None):
        """Set cell properties."""
        cell = ws[coord]
        cell.value = value
        cell.font = font or NORM
        cell.alignment = align
        cell.border = border
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt

    # ── Sheet 1: Rental Income ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Rental Income'
    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 10.09
    ws1.column_dimensions['C'].width = 24
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 11.73

    # Row 1: plain bold title, no fill, left-aligned
    ws1['A1'].value = f'Monthly Rental Income of {month_label}'
    ws1['A1'].font = TITLE
    ws1['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws1['A1'].fill = PatternFill(fill_type=None)
    ws1.row_dimensions[1].height = 22

    # Row 2 blank, Row 3 steel-blue header
    LNW = Alignment(horizontal='left', vertical='center', wrap_text=False)
    for col, lbl in [('A', 'Description'), ('B', 'Amount'),
                     ('C', 'Description'), ('D', 'Amount'), ('E', 'Subtotal')]:
        bdr = BM if col == 'A' else BT
        sc(ws1, f'{col}3', lbl, font=HDR_FONT, fill=STEEL, align=LNW, border=bdr)
    ws1.row_dimensions[3].height = 15

    data_start = 4
    irow = data_start
    for r in ctx['income_rows']:
        if r['is_single']:
            sc(ws1, f'A{irow}', r['left_name'], font=BOLD, border=BM)
            sc(ws1, f'B{irow}', '-', align=R, border=BT)
            sc(ws1, f'C{irow}', '-', font=BOLD, border=BT)
            sc(ws1, f'D{irow}', float(r['left_amount']), align=R, border=BT, fmt='#,##0')
            sc(ws1, f'E{irow}', float(r['left_amount']), align=R, border=BT, fmt='#,##0')
        else:
            sc(ws1, f'A{irow}', r['left_name'], font=BOLD, border=BM)
            sc(ws1, f'B{irow}', float(r['left_amount']), align=R, border=BT, fmt='#,##0')
            sc(ws1, f'C{irow}', r['right_name'], font=BOLD, border=BT)
            sc(ws1, f'D{irow}', float(r['right_amount']), align=R, border=BT, fmt='#,##0')
            sc(ws1, f'E{irow}', f'=B{irow}+D{irow}', align=R, border=BT, fmt='#,##0')
        ws1.row_dimensions[irow].height = 15
        irow += 1

    last_income = irow - 1
    # Total row – light green
    sc(ws1, f'A{irow}', 'Total (Taka)', font=BOLD, fill=LGREEN, border=BM)
    for col in ['B', 'C', 'D']:
        ws1[f'{col}{irow}'].fill = LGREEN
        ws1[f'{col}{irow}'].border = BT
    sc(ws1, f'E{irow}', f'=SUM(E{data_start}:E{last_income})',
       font=BOLD, fill=LGREEN, align=R, border=BT, fmt='#,##0')
    ws1.row_dimensions[irow].height = 15
    income_total_ref = f"'Rental Income'!E{irow}"

    # ── Sheet 2: Expense ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Expense')
    ws2.column_dimensions['A'].width = 42.09
    ws2.column_dimensions['B'].width = 13.0
    ws2.column_dimensions['C'].width = 13.0
    ws2.column_dimensions['D'].width = 63.91

    # Row 1: steel-blue header — left aligned, no wrap (matches original)
    LNW = Alignment(horizontal='left', vertical='center', wrap_text=False)
    for col, lbl in [('A', 'Monthly Expense'), ('B', 'Taka'),
                     ('C', 'Additional'), ('D', 'Comment')]:
        bdr = BM if col == 'A' else BT
        sc(ws2, f'{col}1', lbl, font=HDR_FONT, fill=STEEL, border=bdr, align=LNW)
    ws2.row_dimensions[1].height = 15

    expenses_list = list(ctx['expenses'])
    erow = 2
    amount_refs = []
    for i, exp in enumerate(expenses_list):
        fill = LGREY if i % 2 == 1 else None
        notes = exp.notes or ''
        sc(ws2, f'A{erow}', exp.category.name, fill=fill, border=BM)
        sc(ws2, f'B{erow}', float(exp.amount), fill=fill, align=R, border=BT, fmt='#,##0')
        sc(ws2, f'C{erow}', '', fill=fill, border=BT)
        sc(ws2, f'D{erow}', notes, fill=fill, border=BT,
           align=Alignment(horizontal='left', vertical='top', wrap_text=True))
        if notes:
            ws2.row_dimensions[erow].height = 30
        else:
            ws2.row_dimensions[erow].height = 15
        amount_refs.append(f'B{erow}')
        erow += 1

    # Total row – steel-blue like original
    total_formula = '=' + '+'.join(amount_refs) if amount_refs else '=0'
    sc(ws2, f'A{erow}', 'Total (Taka)', font=HDR_FONT, fill=STEEL, border=BM)
    sc(ws2, f'B{erow}', total_formula, font=HDR_FONT, fill=STEEL, align=R, border=BT, fmt='#,##0')
    for col in ['C', 'D']:
        ws2[f'{col}{erow}'].fill = STEEL
        ws2[f'{col}{erow}'].border = BT
    ws2.row_dimensions[erow].height = 15
    expense_total_ref = f'Expense!B{erow}'
    erow += 2  # blank row gap

    sc(ws2, f'A{erow}', 'Total Maintenance in Account deposit', fill=GREEN_ROW)
    ws2[f'B{erow}'].fill = GREEN_ROW
    ws2[f'C{erow}'].fill = GREEN_ROW
    sc(ws2, f'D{erow}', 'Joint Account', fill=GREEN_ROW)
    ws2.row_dimensions[erow].height = 15
    erow += 1

    sc(ws2, f'A{erow}', 'Total Outstanding Rent', fill=RED_ROW)
    ws2[f'B{erow}'].fill = RED_ROW
    ws2[f'C{erow}'].fill = RED_ROW
    sc(ws2, f'D{erow}', 'Need to be recover', fill=RED_ROW)
    ws2.row_dimensions[erow].height = 15

    # ── Sheet 3: Distribution ────────────────────────────────────────────────
    ws3 = wb.create_sheet('Distribution')
    ws3.column_dimensions['A'].width = 45
    ws3.column_dimensions['B'].width = 14.45
    ws3.column_dimensions['C'].width = 7
    ws3.column_dimensions['D'].width = 55
    ws3.column_dimensions['E'].width = 12

    # Row 1: plain bold title, no fill
    ws3['A1'].value = f'Monthly Distribution of {month_label}'
    ws3['A1'].font = TITLE
    ws3['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws3['A1'].fill = PatternFill(fill_type=None)
    ws3.row_dimensions[1].height = 22

    # Row 3: summary header (row 2 intentionally blank)
    net_formula = f"={income_total_ref}-{expense_total_ref}"
    sc(ws3, 'A3', 'Total Rent after Expense', font=HDR_FONT, fill=DIST_HD, border=BM)
    sc(ws3, 'B3', net_formula, font=HDR_FONT, fill=DIST_HD, align=R, border=BT, fmt='#,##0')
    # C3: (Taka) label — no left border so it reads flush with B
    ws3['C3'].value = '(Taka)'
    ws3['C3'].fill = DIST_HD
    ws3['C3'].border = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    sc(ws3, 'D3', 'Expense per person', font=HDR_FONT, fill=DIST_HD, border=BM)
    sc(ws3, 'E3', 'Amount', font=HDR_FONT, fill=DIST_HD, border=BM)
    ws3.row_dimensions[3].height = 15

    num_owners = len(ctx['dist_rows'])
    drow = 4
    for r in ctx['dist_rows']:
        ded_parts = [f"{d.description}-{int(d.amount):,}" for d in r['deductions']]
        ded_text = ', '.join(ded_parts)
        if r['total_deductions']:
            ded_text += f" = Total {int(r['total_deductions']):,}"

        sc(ws3, f'A{drow}', r['owner'].name, font=BOLD, border=BM)
        sc(ws3, f'B{drow}', f'=B3/{num_owners}', align=R, border=BT, fmt='#,##0')
        # (Taka) label — no left border
        ws3[f'C{drow}'].value = '(Taka)'
        ws3[f'C{drow}'].border = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sc(ws3, f'D{drow}', ded_text or '',
           align=Alignment(horizontal='left', vertical='top', wrap_text=True), border=BT)

        if r['total_deductions']:
            sc(ws3, f'E{drow}', float(r['net']), font=BOLD, align=R, border=BT, fmt='#,##0')
            ws3.row_dimensions[drow].height = 31
        else:
            sc(ws3, f'E{drow}', f'=B{drow}', font=BOLD, align=R, border=BT, fmt='#,##0')
            ws3.row_dimensions[drow].height = 15
        drow += 1

    # ── save & respond ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{month_date.strftime('%B')}_Property_House_{month_date.year}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
